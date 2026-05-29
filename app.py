#!/usr/bin/env python3
"""
Flask主应用
"""
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
from flask_cors import CORS
import pdf_generator
import os
import json
import hashlib
import uuid
from datetime import datetime
from database import Database
import openpyxl
from io import BytesIO

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'finance-receipt-secret-key-2024')

CORS(app)

db = Database()

# ==================== 认证相关 ====================

ADMIN_FILE = os.path.join("data", "admin.json")
SESSION_FILE = os.path.join("data", "sessions.json")


def hash_password(password):
    return hashlib.sha256(password.encode('utf-8')).hexdigest()


def get_admin_config():
    if not os.path.exists(ADMIN_FILE):
        os.makedirs("data", exist_ok=True)
        default_config = {
            "username": "admin",
            "password": hash_password("admin123"),
            "receipt_prefix": "RCP",
            "receipt_seq": 0,
            "last_reset_date": datetime.now().strftime("%Y%m%d")
        }
        with open(ADMIN_FILE, "w", encoding="utf-8") as fp:
            json.dump(default_config, fp, ensure_ascii=False, indent=2)
        return default_config
    with open(ADMIN_FILE, "r", encoding="utf-8") as fp:
        return json.load(fp)


def save_admin_config(config):
    os.makedirs("data", exist_ok=True)
    with open(ADMIN_FILE, "w", encoding="utf-8") as fp:
        json.dump(config, fp, ensure_ascii=False, indent=2)


def get_sessions():
    if not os.path.exists(SESSION_FILE):
        return {}
    with open(SESSION_FILE, "r", encoding="utf-8") as fp:
        return json.load(fp)


def save_sessions(sessions):
    os.makedirs("data", exist_ok=True)
    with open(SESSION_FILE, "w", encoding="utf-8") as fp:
        json.dump(sessions, fp, ensure_ascii=False, indent=2)


def create_session(username):
    token = str(uuid.uuid4())
    sessions = get_sessions()
    sessions[token] = {
        "username": username,
        "created_at": datetime.now().isoformat()
    }
    save_sessions(sessions)
    return token


def validate_session(token):
    if not token:
        return False
    sessions = get_sessions()
    return token in sessions


def destroy_session(token):
    sessions = get_sessions()
    if token in sessions:
        del sessions[token]
        save_sessions(sessions)


def require_auth(f):
    """装饰器：要求管理员登录"""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            token = session.get('admin_token', '')
        if not validate_session(token):
            return jsonify({'error': '请先登录', 'need_login': True}), 401
        return f(*args, **kwargs)
    return decorated


# ==================== 前端页面路由 ====================

@app.route('/')
def index():
    return render_template('dashboard.html')


@app.route('/generate')
def generate():
    return render_template('generate.html')


@app.route('/receipts')
def receipt_list():
    return render_template('receipts.html')


@app.route('/receipt/<int:receipt_id>')
def receipt_detail(receipt_id):
    return render_template('receipt_detail.html', receipt_id=receipt_id)


@app.route('/templates')
def templates_page():
    return render_template('templates.html')


@app.route('/statistics')
def statistics_page():
    return render_template('statistics.html')


@app.route('/batch')
def batch_page():
    return render_template('batch.html')


@app.route('/admin')
def admin_page():
    return render_template('admin.html')


# ==================== 认证API ====================

@app.route('/api/auth/login', methods=['POST'])
def admin_login():
    try:
        data = request.get_json(force=True)
        username = data.get('username', '')
        password = data.get('password', '')

        config = get_admin_config()
        if username == config['username'] and hash_password(password) == config['password']:
            token = create_session(username)
            session['admin_token'] = token
            return jsonify({'success': True, 'token': token, 'username': username})
        return jsonify({'error': '用户名或密码错误'}), 401
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/auth/logout', methods=['POST'])
def admin_logout():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        token = session.get('admin_token', '')
    destroy_session(token)
    session.pop('admin_token', None)
    return jsonify({'success': True})


@app.route('/api/auth/check', methods=['GET'])
def auth_check():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        token = session.get('admin_token', '')
    if validate_session(token):
        config = get_admin_config()
        return jsonify({'logged_in': True, 'username': config['username']})
    return jsonify({'logged_in': False})


@app.route('/api/auth/change-password', methods=['POST'])
@require_auth
def change_password():
    try:
        data = request.get_json(force=True)
        old_password = data.get('old_password', '')
        new_password = data.get('new_password', '')

        config = get_admin_config()
        if hash_password(old_password) != config['password']:
            return jsonify({'error': '原密码错误'}), 400
        if len(new_password) < 6:
            return jsonify({'error': '新密码至少6位'}), 400

        config['password'] = hash_password(new_password)
        save_admin_config(config)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 管理员API ====================

@app.route('/api/admin/receipt-number-config', methods=['GET'])
@require_auth
def get_receipt_number_config():
    config = get_admin_config()
    return jsonify({
        'prefix': config.get('receipt_prefix', 'RCP'),
        'seq': config.get('receipt_seq', 0),
        'last_reset_date': config.get('last_reset_date', '')
    })


@app.route('/api/admin/receipt-number-config', methods=['PUT'])
@require_auth
def update_receipt_number_config():
    try:
        data = request.get_json(force=True)
        config = get_admin_config()

        if 'prefix' in data:
            prefix = data['prefix'].strip().upper()
            if not prefix:
                return jsonify({'error': '编号前缀不能为空'}), 400
            config['receipt_prefix'] = prefix

        if 'reset_seq' in data and data['reset_seq']:
            config['receipt_seq'] = 0
            config['last_reset_date'] = datetime.now().strftime("%Y%m%d")

        if 'set_seq' in data:
            config['receipt_seq'] = int(data['set_seq'])

        save_admin_config(config)
        return jsonify({
            'success': True,
            'prefix': config['receipt_prefix'],
            'seq': config['receipt_seq']
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 收据管理API ====================

@app.route('/api/receipts', methods=['GET'])
def get_receipts():
    try:
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)
        search = request.args.get('search', '')
        currency = request.args.get('currency', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        status = request.args.get('status', '')
        result = db.get_receipts(page=page, limit=limit,
                              search=search if search else None,
                              currency=currency if currency else None,
                              date_from=date_from if date_from else None,
                              date_to=date_to if date_to else None,
                              status=status if status else None)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts', methods=['POST'])
def create_receipt():
    try:
        data = request.get_json(force=True)
        required_fields = ['payer_name', 'amount', 'payment_date', 'purpose', 'payee_name']
        for f in required_fields:
            if f not in data or not data[f]:
                return jsonify({'error': f'缺少必填字段: {f}'}), 400

        if 'receipt_number' not in data or not data['receipt_number']:
            data['receipt_number'] = db.generate_receipt_number()
        amount = float(data['amount'])
        tax_rate = float(data.get('tax_rate', 0))
        tax_amount = amount * tax_rate / 100
        data['tax_amount'] = round(tax_amount, 2)
        data['total_amount'] = round(amount + tax_amount, 2)

        receipt_id = db.create_receipt(data)
        return jsonify({'success': True, 'receipt_id': receipt_id, 'receipt_number': data['receipt_number']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts/<int:receipt_id>', methods=['GET'])
def get_receipt_detail(receipt_id):
    try:
        receipt = db.get_receipt(receipt_id)
        if not receipt:
            return jsonify({'error': '收据不存在'}), 404
        currency_symbols = {'CNY': '¥', 'USD': '$', 'EUR': '€', 'GBP': '£', 'JPY': '¥'}
        receipt['currency_symbol'] = currency_symbols.get(receipt['currency'], '$')
        return jsonify(receipt)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts/<int:receipt_id>', methods=['PUT'])
def update_receipt(receipt_id):
    try:
        data = request.get_json(force=True)
        if 'amount' in data:
            amount = float(data['amount'])
            tax_rate = float(data.get('tax_rate', 0))
            data['tax_amount'] = round(amount * tax_rate / 100, 2)
            data['total_amount'] = round(amount + data['tax_amount'], 2)
        success = db.update_receipt(receipt_id, data)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts/<int:receipt_id>', methods=['DELETE'])
@require_auth
def delete_receipt(receipt_id):
    try:
        success = db.delete_receipt(receipt_id)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts/<int:receipt_id>/pdf', methods=['GET'])
def generate_receipt_pdf(receipt_id):
    try:
        receipt = db.get_receipt(receipt_id)
        if not receipt:
            return jsonify({'error': '收据不存在'}), 404
        template = db.get_default_template()
        template_html = template['template_html'] if template else None

        # 检查模板是否包含不兼容CSS，使用默认模板
        if template_html and ('SimSun' in template_html or 'display: flex' in template_html or '宋体' in template_html):
            template_html = pdf_generator.get_default_template_html()

        pdf_path = pdf_generator.generate_pdf(receipt, template_html)

        # 判断是否内联预览
        inline = request.args.get('inline', '0')
        if inline == '1':
            return send_file(pdf_path, as_attachment=False,
                            mimetype='application/pdf')
        return send_file(pdf_path, as_attachment=True,
                        download_name=f"收据_{receipt['receipt_number']}.pdf",
                        mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts/<int:receipt_id>/preview', methods=['GET'])
def preview_receipt_html(receipt_id):
    """返回收据HTML用于页面内预览"""
    try:
        receipt = db.get_receipt(receipt_id)
        if not receipt:
            return jsonify({'error': '收据不存在'}), 404
        template = db.get_default_template()
        template_html = template['template_html'] if template else None
        html_content = pdf_generator.render_template(template_html or pdf_generator.get_default_template_html(), receipt)
        return html_content
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts/preview-html', methods=['POST'])
def preview_receipt_from_data():
    """根据表单数据返回HTML预览（不保存收据）"""
    try:
        data = request.get_json(force=True)
        # 确保必要字段
        if 'amount' in data:
            amount = float(data['amount'])
            tax_rate = float(data.get('tax_rate', 0))
            data['tax_amount'] = round(amount * tax_rate / 100, 2)
            data['total_amount'] = round(amount + data['tax_amount'], 2)

        template = db.get_default_template()
        template_html = template['template_html'] if template else None
        html_content = pdf_generator.render_template(template_html or pdf_generator.get_default_template_html(), data)
        return html_content
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts/generate-number', methods=['GET'])
def generate_number():
    try:
        return jsonify({'receipt_number': db.generate_receipt_number()})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/receipts/export', methods=['GET'])
def export_receipts():
    """导出收据为Excel文件"""
    try:
        search = request.args.get('search', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        status = request.args.get('status', '')

        result = db.get_receipts(page=1, limit=10000,
                              search=search if search else None,
                              date_from=date_from if date_from else None,
                              date_to=date_to if date_to else None,
                              status=status if status else None)
        receipts = result['receipts']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "收据导出"

        headers = ['收据编号', '付款人', '收款人', '金额', '日期', '状态', '邮箱', '收款事由', '备注']
        ws.append(headers)

        for r in receipts:
            status_text = '已作废' if r.get('status') == 'voided' else '已开具'
            ws.append([
                r.get('receipt_number', ''),
                r.get('payer_name', ''),
                r.get('payee_name', ''),
                r.get('total_amount', 0),
                r.get('payment_date', ''),
                status_text,
                r.get('email', ''),
                r.get('purpose', ''),
                r.get('notes', '')
            ])

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True,
                        download_name=f'收据导出_{datetime.now().strftime("%Y%m%d")}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 付款人/收款人API ====================

@app.route('/api/payers', methods=['GET'])
def get_payers():
    try:
        keyword = request.args.get('keyword', '')
        return jsonify({'payers': db.get_payers(keyword if keyword else None)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/payees', methods=['GET'])
def get_payees():
    try:
        keyword = request.args.get('keyword', '')
        return jsonify({'payees': db.get_payees(keyword if keyword else None)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 批量导入API ====================

@app.route('/api/batch/import', methods=['POST'])
def batch_import():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '请上传Excel文件'}), 400
        file = request.files['file']
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': '请上传.xlsx格式的Excel文件'}), 400

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # 解析表头（第1行）
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val) if val else '')

        # 解析数据行（从第2行开始）
        results = []
        errors = []
        for row_idx in range(2, ws.max_row + 1):
            try:
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    row_data[header] = ws.cell(row=row_idx, column=col_idx).value

                # 提取字段
                payment_date = row_data.get('填制日期', '')
                if isinstance(payment_date, datetime):
                    payment_date = payment_date.strftime('%Y-%m-%d')

                payer_name = row_data.get('付款人名称', '') or row_data.get('名称', '')
                payee_name = row_data.get('收款人名称', '') or row_data.get('收款人', '')
                email = str(row_data.get('邮箱地址', '')) if row_data.get('邮箱地址') else ''

                # 计算各项金额
                items = []
                total = 0
                for key in ['代收代付社保', '代收代付公积金', '代收代付工资',
                           '代收代付个税', '代收代付商险', '代收代付福利', '代收代付残保金', '代收代付其他']:
                    val = row_data.get(key, 0)
                    if val:
                        try:
                            amount = float(val)
                            if amount > 0:
                                items.append({'name': key.replace('代收代付', ''), 'amount': amount})
                                total += amount
                        except:
                            pass

                if not payer_name or total <= 0:
                    continue

                # 生成收据
                receipt_data = {
                    'receipt_number': db.generate_receipt_number(),
                    'payer_name': payer_name,
                    'amount': total,
                    'currency': 'CNY',
                    'payment_date': str(payment_date) if payment_date else datetime.now().strftime('%Y-%m-%d'),
                    'purpose': '、'.join([i['name'] + '：' + str(i['amount']) for i in items]),
                    'payee_name': str(payee_name) if payee_name else '天津俊途人力资源服务有限公司',
                    'tax_rate': 0,
                    'tax_amount': 0,
                    'total_amount': total,
                    'notes': str(row_data.get('备注', '')) if row_data.get('备注') else '',
                    'email': email
                }
                receipt_id = db.create_receipt(receipt_data)
                results.append({'id': receipt_id, 'receipt_number': receipt_data['receipt_number']})
            except Exception as e:
                errors.append({'row': row_idx, 'error': str(e)})

        return jsonify({
            'success': True,
            'created': len(results),
            'receipts': results,
            'errors': errors
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/batch/template', methods=['GET'])
def download_template():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "批量导入模板"

        # 表头
        headers = ['填制日期', '付款人名称', '收款人名称',
                   '代收代付社保', '代收代付公积金', '代收代付工资', '代收代付个税',
                   '代收代付商险', '代收代付福利', '代收代付残保金', '代收代付其他',
                   '合计', '邮箱地址', '备注']
        ws.append(headers)

        # 示例数据
        example = ['2026-05-22', '示例公司', '天津俊途人力资源服务有限公司',
                   '', '', '23399.77', '', '', '', '', '', '23399.77', 'example@email.com', 'C0247CL001171EZ\n2026.1']
        ws.append(example)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True,
                        download_name='批量导入模板修改版.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 模板管理API ====================

@app.route('/api/templates', methods=['GET'])
def get_templates():
    try:
        return jsonify({'templates': db.get_templates()})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates', methods=['POST'])
@require_auth
def create_template():
    try:
        data = request.get_json(force=True)
        if 'name' not in data or not data['name']:
            return jsonify({'error': '模板名称不能为空'}), 400
        if 'template_html' not in data or not data['template_html']:
            return jsonify({'error': '模板内容不能为空'}), 400
        template_id = db.create_template(data)
        return jsonify({'success': True, 'template_id': template_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 统计API ====================

@app.route('/api/statistics/summary', methods=['GET'])
def get_statistics_summary():
    try:
        return jsonify(db.get_statistics_summary())
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 错误及启动 ====================

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': '接口不存在'}), 404
    return render_template('dashboard.html'), 404


@app.errorhandler(500)
def internal_error(e):
    return jsonify({'error': '服务器内部错误'}), 500




# ==================== 模板更新/删除API ====================

@app.route('/api/templates/<int:template_id>', methods=['PUT'])
@require_auth
def update_template(template_id):
    try:
        data = request.get_json(force=True)
        templates = db.get_templates()
        for t in templates:
            if t['id'] == template_id:
                if 'name' in data:
                    t['name'] = data['name']
                if 'description' in data:
                    t['description'] = data['description']
                if 'template_html' in data:
                    t['template_html'] = data['template_html']
                if data.get('is_default'):
                    for tt in templates:
                        tt['is_default'] = (tt['id'] == template_id)
                t['updated_at'] = datetime.now().isoformat()
                db._write_json(db.TEMPLATES_FILE, templates)
                return jsonify({'success': True})
        return jsonify({'error': '模板不存在'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates/<int:template_id>', methods=['DELETE'])
@require_auth
def delete_template(template_id):
    try:
        templates = db.get_templates()
        target = None
        for t in templates:
            if t['id'] == template_id:
                target = t
                break
        if not target:
            return jsonify({'error': '模板不存在'}), 404
        if target.get('is_default'):
            return jsonify({'error': '不能删除默认模板'}), 400
        new_templates = [t for t in templates if t['id'] != template_id]
        db._write_json(db.TEMPLATES_FILE, new_templates)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500





# ==================== 邮件发送API ====================

@app.route('/api/receipts/<int:receipt_id>/send-email', methods=['POST'])
def send_receipt_email(receipt_id):
    """发送收据PDF到指定邮箱"""
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.application import MIMEApplication

        data = request.get_json(force=True)
        email_to = data.get('email', '').strip()
        if not email_to or '@' not in email_to:
            return jsonify({'error': '请输入有效的邮箱地址'}), 400

        receipt = db.get_receipt(receipt_id)
        if not receipt:
            return jsonify({'error': '收据不存在'}), 404

        # 生成PDF
        template = db.get_default_template()
        template_html = template['template_html'] if template else None

        # 检查模板是否包含不兼容CSS，使用默认模板
        if template_html and ('SimSun' in template_html or 'display: flex' in template_html or '宋体' in template_html):
            template_html = pdf_generator.get_default_template_html()

        pdf_path = pdf_generator.generate_pdf(receipt, template_html)

        # 读取邮件配置
        admin_config = get_admin_config()
        smtp_host = os.environ.get('SMTP_HOST', admin_config.get('smtp_host', ''))
        smtp_port = int(os.environ.get('SMTP_PORT', admin_config.get('smtp_port', 465)))  # 默认465(SSL)
        smtp_user = os.environ.get('SMTP_USER', admin_config.get('smtp_user', ''))
        smtp_pass = os.environ.get('SMTP_PASS', admin_config.get('smtp_pass', ''))
        smtp_from = os.environ.get('SMTP_FROM', admin_config.get('smtp_from', smtp_user))
        smtp_use_ssl = smtp_port == 465  # 465端口使用SSL，其他使用STARTTLS

        if not smtp_host or not smtp_user or not smtp_pass:
            return jsonify({'error': '邮件服务未配置，请在系统管理→邮件发送配置中设置邮箱和密码'}), 400

        # 构建邮件
        msg = MIMEMultipart()
        msg['From'] = smtp_from
        msg['To'] = email_to
        msg['Subject'] = f'收据 - {receipt["receipt_number"]}'

        rn = "\n"
        body = "您好，" + rn + rn + f"附件为收据 {receipt['receipt_number']} 的PDF文件。" + rn + rn + f"付款人：{receipt['payer_name']}" + rn + f"金额：¥{receipt['total_amount']:.2f}" + rn + f"日期：{receipt['payment_date']}" + rn + rn + "此邮件由系统自动发送，请勿回复。"
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        with open(pdf_path, 'rb') as f:
            pdf_attach = MIMEApplication(f.read(), _subtype='pdf')
            pdf_attach.add_header('Content-Disposition', 'attachment',
                                  filename=f'收据_{receipt["receipt_number"]}.pdf')
            msg.attach(pdf_attach)

        # 发送（支持SSL和STARTTLS）
        if smtp_use_ssl:
            import smtplib
            server = smtplib.SMTP_SSL(smtp_host, smtp_port)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
        
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_from, [email_to], msg.as_string())
        server.quit()

        # 记录发送邮箱
        db.update_receipt(receipt_id, {'email': email_to})

        return jsonify({'success': True, 'message': f'收据已发送到 {email_to}'})
    except smtplib.SMTPAuthenticationError as e:
        return jsonify({'error': f'SMTP认证失败：请检查邮箱和密码是否正确。腾讯企业邮箱请使用登录密码，不是授权码。详细错误：{str(e)}'}), 500
    except smtplib.SMTPException as e:
        return jsonify({'error': f'邮件发送失败: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'发送失败: {str(e)}'}), 500


# ==================== 收据作废API ====================

@app.route('/api/receipts/<int:receipt_id>/void', methods=['PUT'])
def void_receipt(receipt_id):
    """作废收据"""
    try:
        receipt = db.get_receipt(receipt_id)
        if not receipt:
            return jsonify({'error': '收据不存在'}), 404
        if receipt.get('status') == 'voided':
            return jsonify({'error': '收据已作废'}), 400
        success = db.update_receipt(receipt_id, {'status': 'voided'})
        if success:
            return jsonify({'success': True, 'message': '收据已作废'})
        return jsonify({'error': '作废失败'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== SMTP配置API ====================

@app.route('/api/admin/smtp-config', methods=['GET'])
@require_auth
def get_smtp_config():
    config = get_admin_config()
    return jsonify({
        'smtp_host': config.get('smtp_host', ''),
        'smtp_port': config.get('smtp_port', 587),
        'smtp_user': config.get('smtp_user', ''),
        'smtp_from': config.get('smtp_from', ''),
        'has_password': bool(config.get('smtp_pass', ''))
    })


@app.route('/api/admin/smtp-config', methods=['PUT'])
@require_auth
def update_smtp_config():
    try:
        data = request.get_json(force=True)
        config = get_admin_config()
        if 'smtp_host' in data:
            config['smtp_host'] = data['smtp_host']
        if 'smtp_port' in data:
            config['smtp_port'] = int(data['smtp_port'])
        if 'smtp_user' in data:
            config['smtp_user'] = data['smtp_user']
        if 'smtp_pass' in data and data['smtp_pass']:
            config['smtp_pass'] = data['smtp_pass']
        if 'smtp_from' in data:
            config['smtp_from'] = data['smtp_from']
        save_admin_config(config)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500





@app.route('/api/admin/test-email', methods=['POST'])
@require_auth
def test_send_email():
    """测试邮件发送"""
    try:
        import smtplib
        from email.mime.text import MIMEText

        data = request.get_json(force=True)
        email_to = data.get('email', '').strip()
        if not email_to or '@' not in email_to:
            return jsonify({'error': '请输入有效邮箱'}), 400

        admin_config = get_admin_config()
        smtp_host = admin_config.get('smtp_host', '')
        smtp_port = int(admin_config.get('smtp_port', 465))
        smtp_user = admin_config.get('smtp_user', '')
        smtp_pass = admin_config.get('smtp_pass', '')
        smtp_use_ssl = smtp_port == 465

        if not smtp_host or not smtp_user or not smtp_pass:
            return jsonify({'error': '请先保存完整的邮件配置'}), 400

        msg = MIMEText('这是一封测试邮件，收到说明邮件配置正确。', 'plain', 'utf-8')
        msg['From'] = smtp_user
        msg['To'] = email_to
        msg['Subject'] = '收据系统 - 邮件测试'

        if smtp_use_ssl:
            server = smtplib.SMTP_SSL(smtp_host, smtp_port)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
        
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, [email_to], msg.as_string())
        server.quit()

        return jsonify({'success': True, 'message': '测试邮件已发送'})
    except smtplib.SMTPAuthenticationError as e:
        return jsonify({'error': f'SMTP认证失败：请检查邮箱和密码。腾讯企业邮箱请使用登录密码。详细错误：{str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== 财务章图片上传API ====================

@app.route('/api/admin/seal', methods=['POST'])
@require_auth
def upload_seal():
    """上传财务章图片"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '请选择图片文件'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '请选择图片文件'}), 400
        if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            return jsonify({'error': '只支持 PNG、JPG、JPEG、GIF 格式'}), 400

        base_path = os.path.abspath(os.path.dirname(__file__))
        seal_dir = os.path.join(base_path, 'static')
        os.makedirs(seal_dir, exist_ok=True)
        seal_path = os.path.join(seal_dir, 'seal.png')

        # 保存为 PNG
        from PIL import Image
        img = Image.open(file.stream)
        # 转换RGBA到RGB（如果需要）
        if img.mode in ('RGBA', 'LA', 'P'):
            # 保留透明通道或转换为RGBA
            img = img.convert('RGBA')
        img.save(seal_path, 'PNG')

        return jsonify({'success': True, 'message': '财务章图片已更新'})
    except ImportError:
        # 如果没有 PIL，直接保存
        base_path = os.path.abspath(os.path.dirname(__file__))
        seal_dir = os.path.join(base_path, 'static')
        os.makedirs(seal_dir, exist_ok=True)
        seal_path = os.path.join(seal_dir, 'seal.png')
        file.save(seal_path)
        return jsonify({'success': True, 'message': '财务章图片已更新'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/seal', methods=['DELETE'])
@require_auth
def delete_seal():
    """删除财务章图片"""
    try:
        base_path = os.path.abspath(os.path.dirname(__file__))
        seal_path = os.path.join(base_path, 'static', 'seal.png')
        if os.path.exists(seal_path):
            os.remove(seal_path)
            return jsonify({'success': True, 'message': '财务章图片已删除'})
        return jsonify({'error': '图片不存在'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    os.makedirs('receipts_pdf', exist_ok=True)
    os.makedirs('data', exist_ok=True)
    db.init_default_templates()
    # 初始化管理员配置
    get_admin_config()
    port = int(os.environ.get('PORT', 5000))
    print("="*50)
    print("财务收据生成应用启动成功！")
    print(f"访问地址: http://localhost:{port}")
    print("默认管理员: admin / admin123")
    print("="*50)
    app.run(debug=False, host='0.0.0.0', port=port)
